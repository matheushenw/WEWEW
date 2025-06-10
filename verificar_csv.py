import pandas as pd
import numpy as np
import numpy_financial as npf
from pathlib import Path
from openpyxl import load_workbook

# ❓ Pergunta ao usuário
resposta = input("Deseja remover parcelas liquidadas? (s/n): ").strip().lower()
remover_liquidadas = resposta == "s"

# 📁 Diretórios
pasta_csv = Path("./planilhas")
saida_pasta = pasta_csv / "processados"
saida_pasta.mkdir(parents=True, exist_ok=True)

# Cabeçalhos para colunas calculadas
cabecalho_calculado = [
    "Nova Parcela Nº", "Novo Fluxo de Valor Presente", "Valor Futuro",
    "Valor Presente", "Qtd Parcelas", "Taxa"
]

# Processa .csv e .xlsx
for entrada in pasta_csv.iterdir():
    if entrada.suffix.lower() not in [".csv", ".xlsx", ".xlsm"]:
        continue
    saida = saida_pasta / f"{entrada.stem}_PROCESSADO.csv"

    pessoas = []
    parcelas_por_pessoa = []
    current_pessoa = None
    current_parcelas = []

    if entrada.suffix.lower() == ".csv":
        with open(entrada, "r", encoding="latin1") as f:
            linhas = f.readlines()
    else:
        wb = load_workbook(entrada, data_only=True)
        ws = wb.active
        linhas = [
            ";".join("" if c is None else str(c) for c in row)
            for row in ws.iter_rows(values_only=True)
        ]

    if not linhas:
        continue

    linha_0 = linhas[0]
    dados = linhas[1:]

    for linha in dados:
        partes = linha.strip().split(";")

        # Detecta tipo de linha
        tipo = ""
        if len(partes) > 12 and partes[12] in ["1", "11"]:
            tipo = partes[12]
        elif len(partes) > 6 and partes[6] in ["1", "11"]:
            tipo = partes[6]

        if tipo == "1":
            if current_pessoa:
                pessoas.append(current_pessoa)
                parcelas_por_pessoa.append(current_parcelas)
                current_parcelas = []
            current_pessoa = partes
        elif tipo == "11":
            current_parcelas.append(partes)

    if current_pessoa:
        pessoas.append(current_pessoa)
        parcelas_por_pessoa.append(current_parcelas)

    primeira_vez = True

    with open(saida, "w", encoding="utf-8-sig") as f:
        f.write(linha_0)

    for pessoa, parcelas in zip(pessoas, parcelas_por_pessoa):
        if not parcelas:
            continue

        df = pd.DataFrame(parcelas)

        # Inicializa colunas de cálculo com vazio
        for i in range(6):
            df[i] = ""

        try:
            df[16] = df[16].str.replace(',', '.', regex=False).astype(float)
            # Inclui parcelas com saldo 0 para evitar taxa negativa
            df_abertas = df[df[16] >= 0].copy()
            if df_abertas.empty:
                raise ValueError("Nenhuma parcela aberta válida")

            # PMT (valor da parcela original)
            try:
                PMT = float(df_abertas[12].iloc[0].replace(',', '.'))
            except Exception:
                raise ValueError("PMT inválido")

            VP = df_abertas[16].sum()
            n = len(df_abertas)
            taxa = npf.rate(nper=n, pmt=-PMT, pv=VP, fv=0)

            # Parcelas decrescentes com taxa aplicada (PRICE invertido)
            novo_fluxo = [
                VP * (taxa * (1 + taxa) ** (n - i - 1)) / ((1 + taxa) ** n - 1)
                for i in range(n)
            ]

            for j, idx in enumerate(df_abertas.index):
                df.at[idx, 0] = j + 1
                df.at[idx, 1] = f"{round(novo_fluxo[j], 2):.2f}".replace(".", ",")
                df.at[idx, 2] = f"{round(PMT * n, 2):.2f}".replace(".", ",")
                df.at[idx, 3] = f"{round(VP, 2):.2f}".replace(".", ",")
                df.at[idx, 4] = n
                df.at[idx, 5] = f"{round(taxa, 6):.6f}".replace(".", ",")

        except Exception:
            pass  # Mantém dados vazios nas colunas A-F em caso de erro

        if remover_liquidadas:
            df = df[df[16] != 0]

        with open(saida, "a", encoding="utf-8-sig") as f:
            f.write(";".join(pessoa[:47]) + "\n")

        if len(df.columns) < 47:
            for _ in range(len(df.columns), 47):
                df[len(df.columns)] = ""
        elif len(df.columns) > 47:
            df = df.iloc[:, :47]

        header = cabecalho_calculado + [str(i) for i in range(6, 47)] if primeira_vez else False

        df.iloc[:, :47].to_csv(
            saida,
            mode="a",
            index=False,
            header=header,
            sep=";",
            encoding="utf-8-sig"
        )

        primeira_vez = False

    print(f"✅ Processado: {entrada.name} ➝ {saida.name}")
